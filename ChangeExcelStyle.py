#改变Excel的样式

import openpyxl
from openpyxl.styles import Font
from openpyxl.styles import colors #如果你要用Hex，就用Color这个库

# 打算在一个wb 里面生成5个sheet，分别demo
# 新建一个WB
wb = openpyxl.Workbook()

# Font 默认的字体是Calibri
ws_name_font = "Font_Demo"
ws_font = wb.active
ws_font.title = ws_name_font
# 生成一个字体 例如 斜体24
italic24 = Font(size=24,italic=True)
ws_font["B3"].font = italic24
ws_font["B3"] = 'italic24'

boldRedFont = Font(name='Times New Roman',bold=True,color=colors.RED)
ws_font['A1'].font = boldRedFont
ws_font['A1'] = 'Red Bold Times New Roman'

# Formulas
ws_name_formulas = "Formula_Demo"
ws_formula = wb.create_sheet(title=ws_name_formulas,index=0)
ws_formula['A1'] = 200
ws_formula['A2'] = 300
ws_formula['A3'] = '=SUM(A1:A2)'


# Setting row height and column width
ws_name_size = "Size_Demo"
ws_size = wb.create_sheet(title=ws_name_size,index=0)
# 设置第一行的高度
ws_size.row_dimensions[1].height = 70
# 设置B列的宽度是20
ws_size.column_dimensions['B'].width = 20
ws_size['A1'] = 'Tall Row'
ws_size['B1'] = 'wide column'

# Merging cells
ws_name_mergeCell = "MergeCell_Demo"
ws_mergeCell = wb.create_sheet(title=ws_name_mergeCell,index=0)
ws_mergeCell.merge_cells('A1:D3')
ws_mergeCell['A1'] = 'merge_cells A1 to D3'
ws_mergeCell.merge_cells('C5:D5')
ws_mergeCell['C5'] = 'Two merged cells'

# Unmerging cells
ws_name_unmergeCell = "UnmergeCell_Demo"
# 将上面的sheet拷贝过来
ws_unmerge = wb.copy_worksheet(wb.get_sheet_by_name(ws_name_mergeCell))
ws_unmerge.title = ws_name_unmergeCell
ws_unmerge.unmerge_cells('A1:D3')
ws_unmerge.unmerge_cells('C5:D5')

wb.save(filename='ExcelStyle.xlsx')

# Chart
ws_name_chart = "Chart_Demo"



