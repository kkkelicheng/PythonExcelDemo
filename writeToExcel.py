import openpyxl
from openpyxl.utils import get_column_letter

# 创建一个新的工作簿
wb = openpyxl.Workbook()

# 活跃的,创建wb,应该自带一个active的表单
sheet = wb.active

# 先取3个变量
sheetName_happy2020 = "happy2020"
sheetName_first = "first"
sheetName_middle = "middle"

# change the name of sheet
print(sheet.title)
sheet.title = sheetName_happy2020
print(wb.get_sheet_names())

# 如果你不调用save ，就不会写到硬盘上面的
# wb.save("pyCreatedExcel.xlsx")

# 修改xlsx 的原则： 不改变源文件，重新取一个名字去保存。防止出错，取同名会覆盖。

# 创建其他的表单
# 创建一个name为first sheet的，插在happy2020的前面,假如不指定index，会放在happy2020的后面
wb.create_sheet(index=0,title=sheetName_first)
wb.create_sheet(index=1,title=sheetName_middle)
wb.create_sheet(index=2,title="willRemove")
print(wb.get_sheet_names())

# 删除一个表单
# 首先获取到要删除的表单，2种方式获取，随便用一个，在readExcel中有写
wb.remove_sheet(wb.get_sheet_by_name("willRemove"))

# 如果你不调用save ，就不会写到硬盘上面的
# wb.save("pyCreatedExcel.xlsx")


"""

==========================向cells中写数据==========================

"""

# 向cells中写数据
sh_2020 = wb.get_sheet_by_name(sheetName_happy2020)
# 「赋值形式1」 以cell为单位赋值
sh_2020["A1"] = "Hello Python"
print(sheet["A1"].value)

# 「赋值形式2」 以row为单位赋值
sh_list = wb.get_sheet_by_name(sheetName_first)
rowsData = [
    ['Number','Batch 1','Batch 2'],
    [2,30,35],
    [4,40,35],
    [6,50,35],
    [9,60,35],
    [10,70,35],
    [12,80,35]
]
for rowData in rowsData:
    # 就是依次赋值
    sh_list.append(rowData)

# 「赋值形式3」 用cell的自带函数赋值，方式1的简写，一句话搞定
sh_m = wb.get_sheet_by_name(sheetName_middle)
for row in range(5,30): #5行到29行
    for col in range(15,30): #15列到29列
        sh_m.cell(column=col,row=row,value=get_column_letter(col))

print('sh_m[aa10] = {}'.format(sh_m['AA10'].value))


wb.save(filename = "pyCreatedExcel.xlsx")




