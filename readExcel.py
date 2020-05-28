import openpyxl
import os
import sys
cwd = os.getcwd()
print(cwd)
# 单独运行，获取到当前这个文件路径
argv0 = sys.argv[0]
# 获取到这个路径所在的文件夹 ，官方说dirname是split函数的[0]值
dirname = os.path.dirname(sys.argv[0])
# 改变当前工作路径
os.chdir(dirname)

# 获取当前工作路径下的xlsx文件  workbook 工作簿
wb = openpyxl.load_workbook("example.xlsx")

# 打印一下工作簿下面的 worksheets
for name in wb.sheetnames:
    print(name)
# sheetThree
# sheetOne
# sheetTwo

# 可以新创建一个表单
#sumSheet = wb.create_sheet("sumSheet")
#print("sumSheet" in wb.sheetnames) #True

# 从工作簿获取一个表单 2种方式
sheetThree = wb.get_sheet_by_name("sheetThree")
# sheetOne = wb["sheetOne"]

# 从工作簿获取当前active的表单 
# active_sheet = wb.active

# Excel中可以知道一个单元格的定位是 column + row ,例如 B4
# 获取一个单元格
cell = sheetThree["B2"]
# 获取单元格的属性
print("行 {} , 列 {} , 当前坐标{} ，当前值{}".format(cell.row,cell.column,cell.coordinate,cell.value))

# 获取单元格的另外一种方式
cell2 = sheetThree.cell(row = 2,column = 2)
print("行 {} , 列 {} , 当前坐标{} ，当前值{}".format(cell2.row,cell2.column,cell2.coordinate,cell2.value))

# 获取一整行
#row_2 = sheetThree[2]
#print(row_2) # A2 B2 C2 ...

# 获取一整列
#column_B = sheetThree["B"]
#print(column_B) # B1 B2 ...

# 获取连续的行 获取2,3行
row_range_2to3 = sheetThree[2:3] 
# 遍历2，3行的cell
for row in row_range_2to3:
    for rowCell in row:
        print(rowCell)


# 获取连续的列 获取B,C列
column_range_B2C = sheetThree["B:C"]
for column in  column_range_B2C:
    for columnCell in column:
        print(columnCell)

print("========1:1,2:2===========")

# 获取一个范围的cell (鼠标斜着拉) 1:1 到 2:2
for row in sheetThree.iter_rows(min_row=1,max_row=2,min_col=1,max_col=2):
    for cell in row:
        print(cell)

print("=========A1:B2==========")

# 获取一个范围的cell (鼠标斜着拉) 1:1 到 2:2 第二种方式,这种是行，列
for row in sheetThree["A1:B2"]:
    for cell in row:
        print(cell)

# 获取本sheet的总行 总列
print("本列表是{}行 * {}列".format(sheetThree.max_row,sheetThree.max_column))



print("========改变column的ABCD..成1234===========")
# 这里要用到openpyxl的utils

from openpyxl.utils import get_column_letter,column_index_from_string
# 就是 A <=> 1 B <=> 2 这样的装换。有的时候列会变成多个字母
print("AB是多少列？{}".format(column_index_from_string("AB")))
print("28列对应的字母是多少？{}".format(get_column_letter(28)))



